# backend/app/utils/excel_export.py
import io
import csv
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def generer_excel_rapport(
    titre: str,
    en_tetes: List[str],
    donnees: List[List[Any]],
    sheet_name: str = "Rapport"
) -> bytes:
    """
    Génère un fichier Excel à partir des données fournies.
    
    Args:
        titre: Titre du rapport (placé en première ligne)
        en_tetes: Liste des en-têtes de colonnes
        donnees: Liste des lignes de données
        sheet_name: Nom de la feuille
    
    Returns:
        bytes: Contenu du fichier Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel limite à 31 caractères
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titre
    ws.merge_cells(f'A1:{get_column_letter(len(en_tetes))}1')
    title_cell = ws['A1']
    title_cell.value = titre
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Date de génération
    ws.merge_cells(f'A2:{get_column_letter(len(en_tetes))}2')
    date_cell = ws['A2']
    date_cell.value = f"Généré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    date_cell.font = Font(size=9, italic=True)
    date_cell.alignment = Alignment(horizontal="right")
    
    # En-têtes (ligne 4)
    for col_idx, header in enumerate(en_tetes, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Données
    for row_idx, ligne in enumerate(donnees, 5):
        for col_idx, valeur in enumerate(ligne, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Formatage des valeurs
            if isinstance(valeur, float):
                cell.value = round(valeur, 2)
                cell.number_format = '#,##0.00'
            elif isinstance(valeur, int):
                cell.value = valeur
                cell.number_format = '#,##0'
            elif valeur is None:
                cell.value = ""
            else:
                cell.value = str(valeur)
            
            cell.border = border_style
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustement des largeurs de colonnes
    for col_idx in range(1, len(en_tetes) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in range(4, len(donnees) + 5):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarde
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generer_csv_rapport(
    en_tetes: List[str],
    donnees: List[List[Any]],
    separateur: str = ';'
) -> bytes:
    """
    Génère un fichier CSV à partir des données.
    
    Args:
        en_tetes: Liste des en-têtes
        donnees: Liste des lignes
        separateur: Séparateur de colonnes (; par défaut pour Excel français)
    
    Returns:
        bytes: Contenu du fichier CSV
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=separateur, quoting=csv.QUOTE_MINIMAL)
    
    # Écriture des en-têtes
    writer.writerow(en_tetes)
    
    # Écriture des données
    for ligne in donnees:
        row = []
        for val in ligne:
            if isinstance(val, float):
                row.append(f"{val:.2f}".replace('.', ','))
            elif val is None:
                row.append("")
            else:
                row.append(str(val))
        writer.writerow(row)
    
    # Conversion en bytes avec BOM pour UTF-8 (compatibilité Excel)
    content = buffer.getvalue().encode('utf-8-sig')
    buffer.close()
    
    return content


def generer_excel_rapport_multifeuilles(
    titre: str,
    feuilles: List[Dict[str, Any]]
) -> bytes:
    """
    Génère un fichier Excel avec plusieurs feuilles.
    
    Args:
        titre: Titre global
        feuilles: Liste de dict avec 'nom', 'en_tetes', 'donnees'
    
    Returns:
        bytes: Contenu du fichier Excel
    """
    wb = Workbook()
    wb.remove(wb.active)  # Supprime la feuille par défaut
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for idx, feuille in enumerate(feuilles):
        ws = wb.create_sheet(title=feuille.get('nom', f'Feuille{idx+1}')[:31])
        
        en_tetes = feuille.get('en_tetes', [])
        donnees = feuille.get('donnees', [])
        
        # Date de génération
        ws.merge_cells(f'A1:{get_column_letter(len(en_tetes))}1')
        date_cell = ws['A1']
        date_cell.value = f"{titre} - Généré le {datetime.now().strftime('%d/%m/%Y')}"
        date_cell.font = Font(size=10, italic=True)
        date_cell.alignment = Alignment(horizontal="right")
        
        # En-têtes
        for col_idx, header in enumerate(en_tetes, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Données
        for row_idx, ligne in enumerate(donnees, 4):
            for col_idx, valeur in enumerate(ligne, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(valeur, float):
                    cell.value = round(valeur, 2)
                    cell.number_format = '#,##0.00'
                elif isinstance(valeur, int):
                    cell.value = valeur
                    cell.number_format = '#,##0'
                elif valeur is None:
                    cell.value = ""
                else:
                    cell.value = str(valeur)
                cell.border = border_style
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ajustement des largeurs
        for col_idx in range(1, len(en_tetes) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in range(3, len(donnees) + 4):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()